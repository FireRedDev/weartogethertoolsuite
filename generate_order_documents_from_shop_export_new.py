from tkinter import *
from tkinter import filedialog
from tkinter import messagebox
from tkinter import simpledialog
from openpyxl import load_workbook
import os
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import traceback
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
from pretty_html_table import build_table
import shutil
def openFile():
    file = filedialog.askopenfilename(
        initialdir="C:/Users/MainFrame/Desktop/", 
        title="Open Excel file", 
        filetypes=[("Excel files", ".xlsx .xltx")]
        )
    pathh.insert(END, file)
    orderinformation = simpledialog.askstring("Bestellinformationen", "Bitte gib die Informationen für den Lieferanten ein")
    df = pd.read_excel(file)
    df["Klasse"] = df["Product Variation"].str.split("|",n=4,expand=True)[2].str.replace("Klasse:","")
    df.rename(columns={"Item Name(löschen)" : "Produktname", "Anzahl ":"Anzahl"}, inplace=True)

    pd.set_option('display.max_columns', 100)  # or 1000
    pd.set_option('display.max_rows', 100)  # or 1000
    pd.set_option('display.max_colwidth', 100)
    df = pd.DataFrame(df.values.repeat(df.Anzahl, axis=0), columns=df.columns)
    df.drop(['Anzahl','Product Variation','Bestellnotiz', 'Bestellung Gesamtsumme(löschen)'], axis=1, inplace=True)
    
    t = pd.CategoricalDtype(categories=['XS', 'S','M','L','XL','XXL','XXXL'], ordered=True)
    df['Größe']=pd.Series(df.Größe, dtype=t)
    df.sort_values(by=['Klasse','Produktname','Farbe','Größe'], inplace=True,ignore_index=True)
    #=WENN(UND(I2="Ja";J2="");D2;WENN(I2="Nein";"";WENNFEHLER(RECHTS(J2;LÄNGE(J2)-50);"")))
    df["Individualisierungstext(zählt nur wenn Individualisierung Ja)"] = df.apply(lambda x: x['Input Fields'] if x['Individualisierung']=='Ja' else "", axis=1)
    #df["Individualisierungstext(zählt nur wenn Individualisierung Ja)"] = np.where((~df['Input Fields'].isnull()) & (~df['Individualisierung']== 'Ja') ,df['Nachnahme (Rechnungsadresse)'],"")
    df["Individualisierungstext(zählt nur wenn Individualisierung Ja)"] = df["Individualisierungstext(zählt nur wenn Individualisierung Ja)"].str[50:]
    df["Individualisierungstext(zählt nur wenn Individualisierung Ja)"] = df.apply(lambda x: x['Nachnahme (Rechnungsadresse)'] if pd.isnull(x['Individualisierungstext(zählt nur wenn Individualisierung Ja)']) else x['Individualisierungstext(zählt nur wenn Individualisierung Ja)'], axis=1)
    df["Karton"] = (df.index / 20 + 1).astype(int)
    df.drop(['Input Fields'], axis=1, inplace=True)
    
    df.sort_values(by=['Karton', 'Klasse','Produktname','Farbe','Größe'], inplace=True,ignore_index=True)
    df['Checkbox']='☐'
    df['Unterschrift']=' '
    

    df["Anzahl"]=1
    #df2= df2.pivot_table(index=['Produktname','Größe','Farbe'], 
                # columns='Individualisierung', 
                # margins = True,
                # aggfunc='size', 
                # fill_value=0)
   
    #wb = load_workbook(filename = 'vorlage_bestellliste_shop.xltx')
    #ws = wb["Orders"]
    #bestellungen = ws.tables["Bestellungen"]
    #print(bestellungen)
    
    df.columns = df.columns.astype(str)
    pd.options.display.float_format = '{:,.0f}'.format
    pivottableastable = df.pivot_table(
    index=["Produktname","Farbe","Größe"], values=["Anzahl","Individualisierung"], aggfunc={'Anzahl':len,'Individualisierung':(lambda x:(x=='Ja').sum())}, margins=True, margins_name='Grand Totals')
    
    pivottableastable = pivottableastable.rename(columns={'Individualisierung': 'Anzahl Personalisierungen'})
    
    pivottableaslist = pd.DataFrame(pivottableastable.to_records())
    key = {"Schulpullover": "JH001",
       "Schulshirt": "B&C001",
       }
    pivottableaslist['Produktname-Lieferant']=pivottableaslist['Produktname']
    pivottableaslist.replace({"Produktname-Lieferant": key}, regex=True, inplace = True)
    with pd.ExcelWriter('orderreport.xlsx', engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Orders')
        pivottableastable.to_excel(writer, sheet_name='Übersicht_Tabelle')
        pivottableaslist.to_excel(writer, sheet_name='Übersicht_Liste')
        text_sheet = writer.book.create_sheet(title='Auftragsinformationen')
        text_sheet.cell(column=1, row=1, value=orderinformation)    
    utf8='<head><meta charset="utf-8"></head>'
    html = build_table(df, 'blue_light')
    #pivottableastable=build_table(pivottableastable, 'blue_light')
    pivottableastable= pivottableastable.to_html()
    pivottableaslist=build_table(pivottableaslist, 'blue_light')
    def createpdf(df):
        try:

            #https://stackoverflow.com/questions/32137396/how-do-i-plot-only-a-table-in-matplotlib
            fig, ax =plt.subplots(figsize=(12,4))
            ax.axis('tight')
            ax.axis('off')
            the_table = ax.table(cellText=df.values,colLabels=df.columns,loc='center')
            
            [t.auto_set_font_size(False) for t in [the_table]]
            [t.set_fontsize(8) for t in [the_table]]

            the_table.auto_set_column_width(col=list(range(len(df.columns)))) # Provide integer list of columns to adjust
            #https://stackoverflow.com/questions/4042192/reduce-left-and-right-margins-in-matplotlib-plot
            pp = PdfPages("bestellliste.pdf")
            pp.savefig(fig, bbox_inches='tight')
            pp.close()    
        except:
            print("Exception occurred when creating a pdf")
            traceback.print_exc()    
    createpdf(df)
    try:
        ws.destroy()
    except:
        messagebox.showerror('showerror','Error closing GUI')

    

ws = Tk()
ws.title("Wear Together Toolsuite")
ws.geometry("400x450")
ws['bg']='#fb0'

txtarea = Text(ws, width=40, height=20)
txtarea.pack(pady=20)

pathh = Entry(ws)
pathh.pack(side=LEFT, expand=True, fill=X, padx=20)



Button(
    ws, 
    text="Open File", 
    command=openFile
    ).pack(side=RIGHT, expand=True, fill=X, padx=20)


ws.mainloop()